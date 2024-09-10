import pandas as pd
from fun import create_directory
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

# Define dataset identifiers and test ranges
datasets = ["WV14", "WV25", "WV69"]
test_ranges = ["test_range_1", "test_range_2"]
columns = ["lh", "r2", "rmse", "nrmse", "mae", "nmae"]

# Loop through each test range and dataset
for column in columns:
    for test_range in test_ranges:
        for dataset in datasets:
            # Define the CSV file paths for each condition
            csv_files = [
                Path(f"../final_results/individual_feature_analysis/{column}/{test_range}/final_results_{column}_{test_range}_tar_wo_outliers_w_noise_in_{dataset}.csv"),
                Path(f"../final_results/individual_feature_analysis/{column}/{test_range}/final_results_{column}_{test_range}_tar_wo_outliers_wo_noise_in_{dataset}.csv"),
                Path(f"../final_results/individual_feature_analysis/{column}/{test_range}/final_results_{column}_{test_range}_w_outliers_w_noise_in_{dataset}.csv"),
                Path(f"../final_results/individual_feature_analysis/{column}/{test_range}/final_results_{column}_{test_range}_w_outliers_wo_noise_in_{dataset}.csv")
            ]

            # Create an empty list to store the dataframes for this combination
            dfs = []

            # Loop through each file path and append the DataFrame to the list
            for file in csv_files:
                if file.exists():
                    df = pd.read_csv(file)
                    dfs.append(df)
                else:
                    print(f"File not found: {file}")

            # Concatenate the DataFrames for the current dataset and test_range
            if dfs:
                final_df = pd.concat(dfs, ignore_index=True)

                # Find the maximum values for each column
                max_values = final_df.max(numeric_only=True)

                # Append maximum values as a new row to the DataFrame
                max_row = pd.DataFrame([['Maximum'] + max_values.tolist()], columns=final_df.columns)
                final_df = pd.concat([final_df, max_row], ignore_index=True)

                # Save to Excel
                output_dir = f"../best_results/individual_feature_analysis/{column}/"
                create_directory(output_dir)
                output_file = Path(f"{output_dir}best_{column}_{dataset}_{test_range}.xlsx")
                final_df.to_excel(output_file, index=False)

                # Load the workbook to apply formatting
                workbook = load_workbook(output_file)
                sheet = workbook.active

                # Find and bold the maximum values row
                max_row_index = final_df.shape[0]  # This is the last row index
                for col_num, value in enumerate(final_df.columns, start=1):
                    cell = sheet.cell(row=max_row_index + 1, column=col_num)
                    if cell.value != 'Maximum':  # Skip the label cell
                        cell.font = Font(bold=True)

                # Save the workbook with formatting
                workbook.save(output_file)

                print(f"CSV files for dataset '{dataset}' and test range '{test_range}' have been concatenated into '{output_file}' with maximum values highlighted.")
            else:
                print(f"No CSV files found for dataset '{dataset}' and test range '{test_range}'.")
